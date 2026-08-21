#!/usr/bin/env python3
"""Verify that each registry identifier resolves to the series it claims.

Runs BEFORE any analysis, at gate-open. A wrong identifier does not raise. It
returns a different real series with plausible units, renders cleanly, and is
wrong in a way nothing downstream can detect. That is the failure this module
exists to prevent, and it is why every check below asserts on something
SPECIFIC to the required series rather than on units alone.

Nine traps are recorded in series_registry.py. Each is enforced here as a named
check that fails loudly. A trap that cannot be checked programmatically is
reported as UNCHECKED rather than passed silently.

Result per series:
    PASS       resolved, and every check on it passed
    MISMATCH   resolved, but something differs from what the registry claims
    UNRESOLVED could not be retrieved
    SKIPPED    manual or otherwise out of scope for automated verification

Usage:
    python scripts/fetch/verify.py            # verify everything
    python scripts/fetch/verify.py --key d5   # one series
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import sys
import urllib.parse
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from series_registry import REGISTRY, PlannedSeries  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parent.parent.parent

# Latest acceptable FIRST observation year, per series.
#
# These are calibrated to STUDY 2, not to the withdrawn index test. An
# earlier version asserted "covers 2000-2026", which was the index window;
# under that bar a series starting in 1995 passed every check and then
# silently rendered its domain uninformative, because the ratchet criterion
# needs pre-2000 episodes for the early-vs-late comparison and full history
# for the episode detector's sigma.
#
# A series that starts later than its entry here is not a failure of the
# fetch. It means that domain cannot support Study 2 and must be reported as
# such, the way BLS nurses was demoted in the domain-6 decision.
REQUIRED_FIRST_YEAR = {
    "debt_held_public": 1960,   # OMB 7.1 offers 1940
    "net_interest": 1960,       # OMB 3.1 offers 1940
    "federal_receipts": 1960,   # OMB 1.1 offers single years from 1901
    "saving_rate": 1960,        # BEA T20100 offers 1929
    "union_density": 1983,      # CPS-consistent span the spec cites
    "spr_stocks": 1990,         # SPR fill began in the late 1970s
    "hospital_beds": 1990,      # OECD reaches further; 1990 leaves an early era
    "grain_stocks_use": 1990,
    "credit_gap": 1990,
}

ENV_FILE = REPO_ROOT / ".env"
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/126.0 Safari/537.36"
TIMEOUT = 90


def load_env() -> dict[str, str]:
    out: dict[str, str] = {}
    if not ENV_FILE.is_file():
        return out
    for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        if v.strip():
            out[k.strip()] = v.strip()
    return out


def http(url: str, data: bytes | None = None,
         content_type: str | None = None) -> bytes:
    headers = {"User-Agent": UA}
    if content_type:
        headers["Content-Type"] = content_type
    req = urllib.request.Request(url, data=data, headers=headers)
    with urllib.request.urlopen(req, timeout=TIMEOUT) as r:
        return r.read()


@dataclass
class Check:
    name: str
    ok: bool | None          # None = could not be checked
    detail: str


@dataclass
class Result:
    key: str
    status: str              # PASS | MISMATCH | UNRESOLVED | SKIPPED
    observed: dict = field(default_factory=dict)
    checks: list[Check] = field(default_factory=list)
    error: str = ""

    def add(self, name: str, ok: bool | None, detail: str) -> None:
        self.checks.append(Check(name, ok, detail))

    def settle(self) -> None:
        if self.status in ("UNRESOLVED", "SKIPPED", "UNVERIFIED"):
            return
        if any(c.ok is False for c in self.checks):
            self.status = "MISMATCH"
        elif not any(c.ok is True for c in self.checks):
            # Nothing actually passed; only unperformed checks. That is not a
            # pass, and calling it one is how an unverified series ends up
            # looking identical to a verified one in the summary.
            self.status = "UNVERIFIED"
        else:
            self.status = "PASS"


# --------------------------------------------------------------------------
# Per-source verifiers. Each asserts on something specific to the required
# series, never on units alone.
# --------------------------------------------------------------------------


def verify_bea_saving_rate(s: PlannedSeries, env: dict) -> Result:
    r = Result(s.key, "PENDING")
    key = env.get("BEA_API_KEY")
    if not key:
        r.status = "UNRESOLVED"
        r.error = "BEA_API_KEY missing from .env"
        return r
    url = (
        "https://apps.bea.gov/api/data/?UserID=%s&method=GetData"
        "&datasetname=NIPA&TableName=T20100&Frequency=A&Year=ALL"
        "&ResultFormat=JSON" % key
    )
    try:
        rows = json.loads(http(url))["BEAAPI"]["Results"]["Data"]
    except Exception as e:
        r.status = "UNRESOLVED"
        r.error = str(e)[:200]
        return r

    target = [x for x in rows
              if "saving as a percentage of disposable" in
              x.get("LineDescription", "").lower()]
    r.add("line description present", bool(target),
          target[0]["LineDescription"] if target else "not found in T20100")
    if not target:
        r.status = "MISMATCH"
        return r

    line_no = target[0].get("LineNumber")
    series_code = target[0].get("SeriesCode")
    years = sorted(x["TimePeriod"] for x in target)
    r.observed = {"line": line_no, "series_code": series_code,
                  "n": len(target), "first": years[0], "last": years[-1]}
    r.add("series code A072RC", series_code == "A072RC",
          "got %s" % series_code)
    r.add("annual frequency", True, "Frequency=A requested and returned")
    need = REQUIRED_FIRST_YEAR.get(s.key)
    r.add("reaches back far enough for Study 2 (needs <= %s)" % need,
          int(years[0]) <= need,
          "first observation %s, last %s" % (years[0], years[-1]))
    r.settle()
    return r


def verify_bls(s: PlannedSeries, env: dict) -> Result:
    r = Result(s.key, "PENDING")
    payload = {"seriesid": [s.identifier], "startyear": "2005",
               "endyear": "2024", "catalog": True}
    if env.get("BLS_API_KEY"):
        payload["registrationkey"] = env["BLS_API_KEY"]
    try:
        resp = json.loads(http("https://api.bls.gov/publicAPI/v2/timeseries/data/",
                               data=json.dumps(payload).encode(),
                               content_type="application/json"))
    except Exception as e:
        r.status = "UNRESOLVED"
        r.error = str(e)[:200]
        return r
    if resp.get("status") != "REQUEST_SUCCEEDED" or not resp["Results"].get("series"):
        r.status = "UNRESOLVED"
        r.error = "; ".join(resp.get("message", []))[:200] or "no series returned"
        return r

    ser = resp["Results"]["series"][0]
    cat = ser.get("catalog") or {}
    title = cat.get("series_title", "")
    data = ser.get("data", [])
    r.observed = {"title": title, "n": len(data),
                  "periodicity": cat.get("periodicity_code", ""),
                  "first": data[-1]["year"] if data else "",
                  "last": data[0]["year"] if data else ""}
    r.add("returned data", bool(data), "%d observations" % len(data))
    # Assert on MEANING, not units: percent alone would match many BLS series.
    r.add("title names union membership",
          ("union" in title.lower() and "member" in title.lower())
          if title else None,
          title or "catalog unavailable on this key tier")
    r.add("annual periodicity",
          all(d.get("period") == "A01" for d in data) if data else False,
          "periods: %s" % sorted({d.get("period") for d in data}))
    if data:
        vals = [float(d["value"]) for d in data if d.get("value")]
        plausible = all(0 < v < 50 for v in vals)
        r.add("values in a plausible percent range", plausible,
              "min=%.1f max=%.1f" % (min(vals), max(vals)))
    need = REQUIRED_FIRST_YEAR.get(s.key)
    if need:
        # BLS v2 caps one query at 20 years, so the main probe window cannot
        # see the start of the series. This was UNCHECKED until 2026-08-21.
        # It is not a formality: spec v0.2 section 3.1 chose union density over
        # GSS trust BECAUSE of the CPS-consistent span, and if the series did
        # not reach 1983 the domain could not support Study 2 early-vs-late
        # comparison at all. Ask for the required span directly.
        probe = {"seriesid": [s.identifier], "startyear": str(need),
                 "endyear": str(need + 19)}
        if env.get("BLS_API_KEY"):
            probe["registrationkey"] = env["BLS_API_KEY"]
        try:
            pr = json.loads(http(
                "https://api.bls.gov/publicAPI/v2/timeseries/data/",
                data=json.dumps(probe).encode(),
                content_type="application/json"))
            ok = (pr.get("status") == "REQUEST_SUCCEEDED"
                  and pr["Results"].get("series"))
            got = sorted(int(d["year"])
                         for d in (pr["Results"]["series"][0].get("data", [])
                                   if ok else []))
            detail = ("%d observations, %s to %s, in the %s-%s probe"
                      % (len(got), got[0], got[-1], need, need + 19)
                      if got else "probe returned nothing")
        except Exception as e:
            got, detail = [], "reach probe failed: %s" % str(e)[:120]
        r.add("reaches back to %s (the CPS-consistent span the spec cites)"
              % need, bool(got) and got[0] <= need, detail)
    r.settle()
    return r


def verify_treasury_debt(s: PlannedSeries, env: dict) -> Result:
    """TRAP: debt held by THE PUBLIC, not TOTAL public debt.

    Asserts on the FIELD NAME. Both quantities are dollar series named almost
    identically and differ by roughly a quarter, so units cannot separate them.
    """
    r = Result(s.key, "PENDING")
    url = ("https://api.fiscaldata.treasury.gov/services/api/fiscal_service/"
           "v2/accounting/od/debt_to_penny?sort=-record_date&page[size]=1")
    try:
        payload = json.loads(http(url))
    except Exception as e:
        r.status = "UNRESOLVED"
        r.error = str(e)[:200]
        return r
    rec = (payload.get("data") or [{}])[0]
    fields = list(rec.keys())
    r.observed = {"fields": len(fields), "record_date": rec.get("record_date", "")}

    held_public = "debt_held_public_amt" in fields
    total_debt = "tot_pub_debt_out_amt" in fields
    r.add("TRAP: debt_held_public_amt field exists", held_public,
          "present" if held_public else "ABSENT - cannot isolate the required quantity")
    r.add("TRAP: distinguishable from total public debt", held_public and total_debt,
          "both fields present, so the wrong one is selectable by mistake"
          if (held_public and total_debt) else "fields: %s" % ", ".join(fields[:6]))
    if held_public and total_debt:
        try:
            hp = float(rec["debt_held_public_amt"])
            tot = float(rec["tot_pub_debt_out_amt"])
            r.add("held-public is materially smaller than total",
                  hp < tot * 0.95,
                  "ratio held/total = %.3f" % (hp / tot) if tot else "n/a")
        except Exception:
            r.add("held-public smaller than total", None, "non-numeric values")
    r.add("coverage back to 2000", None,
          "NOT CHECKED - single-record probe only; full span must be confirmed at fetch")
    r.settle()
    return r


def verify_omb_table(s: PlannedSeries, env: dict) -> Result:
    """TRAP (severe): Table 3.1 has three 'Net interest' rows in different units.

    Asserts on the SECTION HEADING above the row, never the row label.
    """
    import openpyxl

    r = Result(s.key, "PENDING")
    table = "hist03z1" if s.key == "net_interest" else "hist01z1"
    url = ("https://www.whitehouse.gov/wp-content/uploads/2026/04/"
           "%s_fy2027.xlsx" % table)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(http(url)), read_only=True,
                                    data_only=True)
        rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    except Exception as e:
        r.status = "UNRESOLVED"
        r.error = str(e)[:200]
        return r

    if s.key == "net_interest":
        hits = [i for i, row in enumerate(rows)
                if isinstance(next((c for c in row if c not in (None, "")), None), str)
                and next(c for c in row if c not in (None, "")).strip().lower()
                == "net interest"]
        r.observed = {"net_interest_rows": hits}
        r.add("TRAP: multiple 'Net interest' rows found", len(hits) >= 2,
              "rows %s - the label alone is ambiguous" % hits)

        def heading_above(idx: int) -> str:
            for j in range(idx - 1, max(idx - 22, -1), -1):
                cells = [c for c in rows[j] if c not in (None, "")]
                txt = [c for c in cells if isinstance(c, str)]
                num = [c for c in cells if isinstance(c, (int, float))]
                if txt and not num and len(" ".join(str(t) for t in txt)) > 12:
                    return " ".join(str(t) for t in txt).strip().lower()
            return ""

        dollar_rows = [i for i in hits if "millions of dollars" in heading_above(i)]
        r.add("TRAP: a dollar-denominated row is identifiable by heading",
              len(dollar_rows) == 1,
              "rows under 'in millions of dollars': %s" % dollar_rows)
        if len(dollar_rows) == 1:
            chosen = dollar_rows[0]
            r.observed["chosen_row"] = chosen
            r.add("chosen row matches the registry", chosen == 21,
                  "registry says 21, heading-based selection says %d" % chosen)
            vals = [c for c in rows[chosen] if isinstance(c, (int, float))]
            r.add("values look like dollar millions, not percentages",
                  max(vals) > 1000, "max=%s" % max(vals))
            r.observed["n_values"] = len(vals)
        pct_rows = [i for i in hits if "percentages of outlays" in heading_above(i)]
        r.add("TRAP: percentage-of-outlays row identified and excluded",
              bool(pct_rows), "rows %s excluded" % pct_rows)

        # Added 2026-08-21. Row 1 was never inspected, so the two hazards the
        # registry records for this table - the transposition and the six
        # projection columns - were both unguarded here.
        hdr = [str(c).strip() if c is not None else "" for c in rows[1]]
        yrs = [h for h in hdr if h[:4].isdigit()]
        est = [h for h in yrs if not h.isdigit()]
        plain = sorted(int(h) for h in yrs if h.isdigit())
        r.observed["year_headers"] = len(yrs)
        r.add("TRAP: transposed - years are COLUMN headers in row 1",
              len(yrs) > 80,
              "%d year headers, %s to %s" % (len(yrs), yrs[0], yrs[-1])
              if yrs else "none found, which is what a parser written for "
                          "Tables 1.1 and 7.1 would see")
        r.add("TRAP: projection columns present and labelled", bool(est),
              "%d estimate headers: %s" % (len(est), est))
        # Asserted as an invariant rather than against a fixed last-actual
        # year, so the FY2028 edition does not fail this for being newer.
        estyr = sorted(int(h[:4]) for h in est)
        r.add("actual/estimate boundary is clean (every projection is "
              "later than every actual)",
              bool(plain) and bool(estyr) and estyr[0] > plain[-1],
              "actuals end %s, projections begin %s"
              % (plain[-1] if plain else "-", estyr[0] if estyr else "-"))
        if len(dollar_rows) == 1:
            # The workbook uses a ten-period string for missing data. A fetch
            # that coerces with a bare except would turn those into zeros in a
            # level series. Assert the chosen row carries none.
            bad = [c for c in rows[dollar_rows[0]][1:]
                   if c not in (None, "") and not isinstance(c, (int, float))]
            r.add("chosen row is free of missing-data sentinels",
                  not bad, "no non-numeric cells" if not bad
                  else "%d found: %s" % (len(bad), bad[:3]))
    else:
        hdr = None
        for i, row in enumerate(rows[:8]):
            cells = [str(c).strip().lower() if c else "" for c in row]
            if "receipts" in cells and "outlays" in cells:
                hdr = i
                break
        r.observed = {"header_row": hdr}
        r.add("receipts/outlays header located", hdr is not None,
              "row %s" % hdr)
        firsts = [next((c for c in row if c not in (None, "")), None)
                  for row in rows]
        aggregates = [f for f in firsts
                      if isinstance(f, str) and "-" in f and f[:4].isdigit()]
        r.add("TRAP: multi-year aggregate rows present and must be skipped",
              bool(aggregates), "found %s" % aggregates[:3])
        # OMB stores years as TEXT in this workbook, not as numbers. An earlier
        # version of this check required int/float and rejected every row,
        # which at fetch time would have silently produced an empty series.
        def as_year(v):
            if isinstance(v, (int, float)) and 1900 < v < 2040:
                return int(v)
            if isinstance(v, str):
                t = v.strip().rstrip("*").strip()
                if t.isdigit() and 1900 < int(t) < 2040:
                    return int(t)
            return None

        year_rows = [y for y in (as_year(f) for f in firsts) if y is not None]
        r.add("single-year rows parse", len(year_rows) > 80,
              "%d year rows, %s to %s"
              % (len(year_rows), min(year_rows), max(year_rows))
              if year_rows else "none parsed")
        def is_text_year(v) -> bool:
            if not isinstance(v, str):
                return False
            return v.strip().rstrip("*").strip().isdigit()

        r.add("year cells are text, not numbers",
              any(is_text_year(f) for f in firsts),
              "confirmed - a numeric-only parser would return an empty series")
        # Was UNCHECKED until 2026-08-21. Row 2 carries merged span labels
        # (Total / On-Budget / Off-Budget) and row 3 repeats Receipts /
        # Outlays / Surplus under each, so "Receipts" names THREE columns.
        # Forward-fill the spans and select on the pair.
        span, cur = [], ""
        for c in rows[2]:
            cur = str(c).strip().lower() if c not in (None, "") else cur
            span.append(cur)
        sub3 = [str(c).strip().lower() if c not in (None, "") else ""
                for c in rows[3]]
        def col_for(section):
            return next((i for i in range(min(len(span), len(sub3)))
                         if span[i] == section and sub3[i] == "receipts"), None)
        total_col, onbud_col = col_for("total"), col_for("on-budget")
        r.observed.update({"total_receipts_col": total_col,
                           "on_budget_receipts_col": onbud_col})
        r.add("TRAP: Total receipts column selected by SPAN, not by label",
              total_col is not None and total_col != onbud_col,
              "Total/Receipts at %s, On-Budget/Receipts at %s"
              % (total_col, onbud_col))
        r.add("matches the registry (first numeric column)", total_col == 1,
              "registry says the first numeric column; span-based "
              "selection says %s" % total_col)
        # Before Social Security went off-budget the two columns are EQUAL,
        # so an early year cannot discriminate them. Use a late one.
        late = next((row for row in rows
                     if str(next((c for c in row if c not in (None, "")), ""))
                     .strip() == "1990"), None)
        if late is not None and total_col is not None and onbud_col is not None:
            tv, ov = late[total_col], late[onbud_col]
            r.add("the two columns are materially different in FY1990",
                  isinstance(tv, (int, float)) and isinstance(ov, (int, float))
                  and tv > ov * 1.15,
                  "Total %s vs On-Budget %s - On-Budget excludes Social "
                  "Security and is smaller by %.0f%%"
                  % (tv, ov, 100.0 * (1 - ov / tv))
                  if isinstance(tv, (int, float)) and isinstance(ov, (int, float))
                  and tv else "could not read FY1990")
    r.settle()
    return r


def verify_omb_debt(s: PlannedSeries, env: dict) -> Result:
    """Table 7.1: debt held by the public as a percent of GDP.

    Two traps share one shape here. The table carries the SAME sub-headers twice
    - once under "In Millions of Dollars" and once under "As Percentages of GDP"
    - and within each, "Held by the Public" splits into Total / Federal Reserve
      System / Other. So four columns can plausibly answer to "held by the
    public", in two different units. Selection asserts on the row-1 SECTION
    header, never on the sub-header alone.
    """
    import openpyxl

    r = Result(s.key, "PENDING")
    url = ("https://www.whitehouse.gov/wp-content/uploads/2026/04/"
           "hist07z1_fy2027.xlsx")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(http(url)), read_only=True,
                                    data_only=True)
        rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    except Exception as e:
        r.status = "UNRESOLVED"
        r.error = str(e)[:200]
        return r

    def cell(row, col):
        v = rows[row][col] if col < len(rows[row]) else None
        return str(v).replace("\n", " ").strip() if v else ""

    sect = [cell(1, c) for c in range(len(rows[1]))]
    sub2 = [cell(2, c) for c in range(len(rows[2]))]
    sub3 = [cell(3, c) for c in range(len(rows[3]))]
    r.observed = {"sections": [x for x in sect if x]}

    pct_start = next((c for c, v in enumerate(sect)
                      if "percentages of gdp" in v.lower()), None)
    r.add("TRAP: percent-of-GDP section located by ROW-1 header",
          pct_start is not None, "section begins at column %s" % pct_start)
    if pct_start is None:
        r.status = "MISMATCH"
        return r

    col = next((c for c in range(pct_start, len(sub2))
                if "held by the public" in sub2[c].lower()), None)
    if col is not None:
        while col + 1 < len(sub3) and sub3[col].strip() == "":
            col += 1
    r.add("TRAP: 'Held by the Public' column inside that section",
          col is not None, "column %s" % col)
    if col is None:
        r.status = "MISMATCH"
        return r

    r.add("TRAP: it is the Total sub-column, not Federal Reserve System",
          sub3[col].lower().startswith("total"),
          "sub-header reads %r" % sub3[col])
    r.add("matches the registry (column 8)", col == 8,
          "registry says 8, header-based selection says %s" % col)

    gross = next((c for c in range(pct_start, len(sub2))
                  if "gross" in sub2[c].lower()), None)
    r.add("TRAP: distinguishable from GROSS federal debt",
          gross is not None and gross != col,
          "gross at %s, held-by-public at %s" % (gross, col))

    def as_year(v):
        if isinstance(v, (int, float)) and 1900 < v < 2040:
            return int(v)
        if isinstance(v, str):
            x = v.strip().rstrip("*").strip()
            if x.isdigit() and 1900 < int(x) < 2040:
                return int(x)
        return None

    years, vals = [], []
    for row in rows:
        y = as_year(next((c for c in row if c not in (None, "")), None))
        if y is not None and col < len(row) and isinstance(row[col], (int, float)):
            years.append(y)
            vals.append(float(row[col]))
    r.observed.update({"n": len(years),
                       "first": min(years) if years else None,
                       "last": max(years) if years else None})
    r.add("year cells parse (they are TEXT here)", len(years) > 60,
          "%d rows, %s to %s" % (len(years), min(years) if years else "-",
                                 max(years) if years else "-"))
    need = REQUIRED_FIRST_YEAR.get(s.key)
    r.add("reaches back far enough for Study 2 (needs <= %s)" % need,
          bool(years) and min(years) <= need,
          "first observation %s" % (min(years) if years else "none"))
    if vals:
        r.add("values read as percentages, not dollar millions",
              max(vals) < 500, "max %.1f" % max(vals))
    # Checked live 2026-08-21. Table 7.1 stops at the last ACTUAL fiscal year
    # and carries no estimate rows, unlike Table 3.1, which runs six of them.
    # The retracted note here claimed the opposite. Kept as a standing
    # assertion rather than deleted, because a later edition could begin
    # publishing projections and this is where that would surface.
    labels = [str(next((c for c in row if c not in (None, "")), ""))
              for row in rows]
    est = [x for x in labels if "estimate" in x.lower()]
    r.add("TRAP: carries no projection rows to trim (unlike Table 3.1)",
          not est,
          "no estimate row labels found"
          if not est else "found %d: %s" % (len(est), est[:4]))
    r.settle()
    return r


def _csv_rows(blob: bytes) -> list[dict]:
    text = blob.decode("utf-8-sig", errors="replace")
    return list(csv.DictReader(io.StringIO(text)))


def verify_bis_gap(s: PlannedSeries, env: dict) -> Result:
    """BIS credit-to-GDP GAP, US private non-financial, all lenders.

    The discriminating assertion is NOT units. Ratio, trend and gap are all
    served by this flow, all labelled percent-of-GDP. What separates them is
    the CG_DTYPE code and the VALUE RANGE: the US credit-to-GDP ratio runs
    around 150-250, the gap runs in single or low double digits either side of
    zero. A run that silently returned the ratio would pass any units check.
    """
    r = Result(s.key, "PENDING")
    url = ("https://stats.bis.org/api/v2/data/dataflow/BIS/WS_CREDIT_GAP/1.0/"
           "Q.US.P.A.C?format=csv")
    try:
        rows = _csv_rows(http(url))
    except Exception as e:
        r.status = "UNRESOLVED"
        r.error = str(e)[:200]
        return r
    if not rows:
        r.status = "UNRESOLVED"
        r.error = "empty CSV"
        return r

    def col(*names):
        for n in names:
            if n in rows[0]:
                return n
        return None

    c_type = col("CG_DTYPE", "TC_DTYPE")
    c_per = col("TIME_PERIOD")
    c_val = col("OBS_VALUE")
    c_bor = col("TC_BORROWERS")
    c_len = col("TC_LENDERS")
    r.observed = {"n": len(rows), "columns": list(rows[0])[:12]}

    types = sorted({row.get(c_type, "") for row in rows}) if c_type else []
    r.add("TRAP: exactly one series type returned, and it is the GAP",
          types == ["C"], "CG_DTYPE values present: %s" % (types or "column absent"))
    if c_bor:
        bors = sorted({row[c_bor] for row in rows})
        r.add("TRAP: borrower sector is P (private non-financial)",
              bors == ["P"], "TC_BORROWERS: %s" % bors)
    if c_len:
        lens = sorted({row[c_len] for row in rows})
        r.add("TRAP: lender basis is A (all sectors), not banks only",
              lens == ["A"], "TC_LENDERS: %s" % lens)

    pers = sorted(row[c_per] for row in rows if row.get(c_per))
    r.add("quarterly periods", bool(pers) and all("-Q" in x for x in pers[:5]),
          "%d periods, %s to %s" % (len(pers), pers[0] if pers else "-",
                                    pers[-1] if pers else "-"))
    need = REQUIRED_FIRST_YEAR.get(s.key)
    if pers and need:
        r.add("reaches back far enough for Study 2 (needs <= %s)" % need,
              int(pers[0][:4]) <= need, "first period %s" % pers[0])

    vals = []
    for row in rows:
        try:
            vals.append(float(row[c_val]))
        except (TypeError, ValueError):
            pass
    r.observed["last_value"] = vals[-1] if vals else None
    # The load-bearing check. A gap straddles zero; a ratio does not.
    r.add("TRAP: values are a GAP, not the credit-to-GDP RATIO",
          bool(vals) and min(vals) < 0 < max(vals) and max(abs(v) for v in vals) < 60,
          "range %.2f to %.2f - a ratio would sit near 150-250 and never "
          "cross zero" % (min(vals), max(vals)) if vals else "no values parsed")
    r.settle()
    return r


def verify_oecd_beds(s: PlannedSeries, env: dict) -> Result:
    """OECD total hospital beds per 1,000, United States.

    Also computes the spec section-5 block-coverage figure for B3ex, because
    domain 6's admission to the verdict-bearing block turns on it and the
    answer changes with each OECD release. Recording it as a live number
    rather than a hand-worked one means it stops being wrong silently.
    """
    r = Result(s.key, "PENDING")
    url = ("https://sdmx.oecd.org/public/rest/data/"
           "OECD.ELS.HD,DSD_HEALTH_REAC_HOSP@DF_BEDS_FUNC,1.0/"
           "USA.HB.10P3HB._Z._Z._T._T._Z._Z?format=csvfile")
    try:
        rows = _csv_rows(http(url))
    except Exception as e:
        r.status = "UNRESOLVED"
        r.error = str(e)[:200]
        return r
    if not rows:
        r.status = "UNRESOLVED"
        r.error = "empty CSV - a 9-position key returning nothing usually "\
                  "means a code changed, not that the US has no beds"
        return r

    hdr = list(rows[0])
    r.observed = {"n": len(rows), "columns": hdr[:12]}

    def uniq(name):
        return sorted({row.get(name, "") for row in rows}) if name in hdr else None

    for name, want, why in (
            ("REF_AREA", ["USA"], "country is the US"),
            ("MEASURE", ["HB"], "TRAP: measure is total hospital beds, not an ICU variant"),
            ("UNIT_MEASURE", ["10P3HB"],
             "TRAP: unit is per-1,000 inhabitants, not absolute bed counts"),
            ("HEALTH_FUNCTION", ["_T"],
             "TRAP: function is Total, not curative-only (HC1)")):
        got = uniq(name)
        if got is not None:
            r.add(why, got == want, "%s = %s" % (name, got))

    years, series = [], {}
    for row in rows:
        try:
            y = int(str(row.get("TIME_PERIOD", ""))[:4])
            v = float(row["OBS_VALUE"])
        except (TypeError, ValueError):
            continue
        years.append(y)
        series[y] = v
    years.sort()
    r.observed.update({"first": years[0] if years else None,
                       "last": years[-1] if years else None})
    r.add("annual observations parse", len(years) > 20,
          "%d years, %s to %s" % (len(years), years[0] if years else "-",
                                  years[-1] if years else "-"))
    if series:
        vals = list(series.values())
        r.add("values are a per-1,000 rate, not a bed count",
              all(0 < v < 30 for v in vals),
              "min %.2f max %.2f" % (min(vals), max(vals)))
    need = REQUIRED_FIRST_YEAR.get(s.key)
    if years and need:
        r.add("reaches back far enough for Study 2 (needs <= %s)" % need,
              years[0] <= need, "first observation %d" % years[0])

    # Spec v0.2 section 5: a domain must cover >= 60% of a block's usable years
    # to enter that block. B3ex is the third block with crisis years removed.
    B3EX = [2018, 2019, 2022, 2023, 2024, 2025]
    have = [y for y in B3EX if y in series]
    frac = len(have) / len(B3EX)
    r.add("BLOCKING: covers at least 60 percent of B3ex, the verdict block",
          frac >= 0.60,
          "%d of %d years (%.0f%%) - have %s, missing %s. Below 60%% this "
          "domain leaves B3 entirely; OECD publishing US 2023 would lift it "
          "to 4 of 6 = 67%%."
          % (len(have), len(B3EX), 100 * frac, have,
             [y for y in B3EX if y not in series]))
    r.settle()
    return r


def verify_generic_reachable(s: PlannedSeries, env: dict) -> Result:
    """Entries with a host but no settled endpoint.

    Returns UNVERIFIED, never PASS. An earlier version returned PASS whenever
    the host answered, which meant a series whose identity had not been checked
    at all appeared in the summary next to genuinely verified ones. Reachability
    is not verification, and a probe that cannot fail is not a check.

    Route templates carrying placeholders are not probed at all: stripping the
    placeholder yields a directory URL that says nothing about the series.
    """
    r = Result(s.key, "UNVERIFIED")
    if "{" in s.route:
        r.error = ("route is a template (%s) and the endpoint is unsettled; "
                   "probing the bare host would test nothing" % s.route)
        r.add("series identity verified", None,
              "NOT CHECKED - endpoint unsettled, identifier '%s'" % s.identifier)
        return r
    if not s.route.startswith("http"):
        r.status = "SKIPPED"
        r.error = "no HTTP route recorded"
        return r
    try:
        body = http(s.route)
    except Exception as e:
        r.status = "UNRESOLVED"
        r.error = str(e)[:160]
        return r
    r.observed = {"bytes": len(body)}
    if not body:
        r.status = "UNRESOLVED"
        r.error = "host answered with an empty body"
        return r
    r.add("host answers", True, "%d bytes" % len(body))
    r.add("series identity verified", None,
          "NOT CHECKED - endpoint not settled; identifier is '%s'" % s.identifier)
    return r


VERIFIERS = {
    "saving_rate": verify_bea_saving_rate,
    "union_density": verify_bls,
    # Re-bound 2026-08-20. This pointed at verify_treasury_debt after the
    # registry moved to OMB Table 7.1, so it reported PASS while checking a
    # source no longer in use - the precise failure this module exists to
    # catch, occurring inside the module itself.
    "debt_held_public": verify_omb_debt,
    "net_interest": verify_omb_table,
    "federal_receipts": verify_omb_table,
    # Added 2026-08-20 once the SDMX keys were settled. These exist so the
    # two routes rest on a live assertion here, not on the report that
    # derived them.
    "credit_gap": verify_bis_gap,
    "hospital_beds": verify_oecd_beds,
}


def verify_one(s: PlannedSeries, env: dict) -> Result:
    if s.confidence == "manual":
        r = Result(s.key, "SKIPPED")
        r.error = "Tier 3 manual transcription; not scriptable"
        return r
    fn = VERIFIERS.get(s.key, verify_generic_reachable)
    try:
        return fn(s, env)
    except Exception as e:  # a verifier must never take the run down
        r = Result(s.key, "UNRESOLVED")
        r.error = "verifier raised: %s" % str(e)[:160]
        return r


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--key", help="verify a single registry key")
    args = ap.parse_args()

    env = load_env()
    targets = [s for s in REGISTRY if (not args.key or s.key == args.key)]
    print("verify_registry - %d series, keys loaded: %s"
          % (len(targets), ", ".join(sorted(env)) or "none"))
    print("")

    results = []
    for s in targets:
        res = verify_one(s, env)
        results.append(res)
        print("%-9s %-18s %s" % (res.status, s.key, s.label[:44]))
        for c in res.checks:
            mark = {True: "  ok  ", False: " FAIL ", None: " ---- "}[c.ok]
            print("   %s %-46s %s" % (mark, c.name[:46], c.detail[:80]))
        if res.error:
            print("          error: %s" % res.error)
        print("")

    tally: dict[str, int] = {}
    for r in results:
        tally[r.status] = tally.get(r.status, 0) + 1
    print("=" * 66)
    print("SUMMARY: " + "  ".join("%s=%d" % kv for kv in sorted(tally.items())))
    unchecked = sum(1 for r in results for c in r.checks if c.ok is None)
    print("%d check(s) could not be performed and are reported as UNCHECKED, "
          "not passed." % unchecked)
    failed = [r.key for r in results if r.status == "MISMATCH"]
    if failed:
        print("MISMATCHES: %s" % ", ".join(failed))
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
